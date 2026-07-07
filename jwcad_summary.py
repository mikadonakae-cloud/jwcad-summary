"""
JwCAD 配線・配管集計ツール
JWWファイルまたは画像から電線・配管情報を抽出して集計表を生成します。
"""

import re
import sys
import struct
import os
from pathlib import Path
from dataclasses import dataclass, field
from collections import defaultdict
from typing import Optional


# ─────────────────────────────────────────
#  データ構造
# ─────────────────────────────────────────

@dataclass
class ConduitEntry:
    """配管1区間の情報"""
    install_type: str          # 露出 / 埋設
    conduit_type: str          # HIVE70 / FEP(80) / FEP(30) など
    length: int                # 長さ (m)
    cables: list[str] = field(default_factory=list)  # 入っているケーブル種

@dataclass
class CableEntry:
    """ケーブル1区間の情報"""
    cable_type: str            # 6.6kVCVT38sq / CVT150sq / IV38sq など
    length: int                # 長さ (m)

@dataclass
class FreeCableEntry:
    """配管なしケーブル（充電器内・ポール内・キュービクル内など）"""
    location: str              # 設置場所（充電器内、ポール内 など）
    cable_type: str
    length: int


# ─────────────────────────────────────────
#  JWWバイナリからテキスト抽出
# ─────────────────────────────────────────

def extract_text_from_jww(filepath: str) -> list[str]:
    """
    JWW v7 バイナリパーサー。
    テキスト要素は \\xff\\xfe\\xff + uint8長さ + UTF-16-LE 形式で格納。
    フォント名＋テキスト本文のペアを抽出し、座標付きで返す。
    座標は content_end+17 バイト目から 8バイト倍精度浮動小数点×2。
    印刷範囲（二重枠）内のテキストのみ対象とするため極端な座標を除外。
    """
    _JWW_MARKER = bytes([0xff, 0xfe, 0xff])
    _FONT_RE = re.compile(r'ＭＳ|Times New Roman|Arial|明朝|Gothic|Serif|Sans')

    try:
        with open(filepath, "rb") as f:
            data = f.read()

        # ── UTF-16 文字列を全抽出 ──
        raw_strings: list[tuple[int, str]] = []  # (offset, text)
        i = 0
        while i < len(data) - 4:
            if data[i:i+3] == _JWW_MARKER:
                length = data[i + 3]
                text_start = i + 4
                text_end = text_start + length * 2
                if text_end <= len(data) and 1 <= length <= 255:
                    try:
                        text = data[text_start:text_end].decode("utf-16-le", errors="strict")
                        raw_strings.append((i, text))
                        i = text_end
                        continue
                    except (UnicodeDecodeError, ValueError):
                        pass
            i += 1

        # ── フォント名＋テキスト本文のペアを抽出し、座標も取得 ──
        # 構造: [font_BOM+len+text] [content_BOM+len+text] [17byte suffix] [x: f64] [y: f64]
        records: list[tuple[float, float, int, str]] = []  # (x, y, offset, text)

        j = 0
        while j < len(raw_strings) - 1:
            off1, font = raw_strings[j]
            off2, content = raw_strings[j + 1]

            if _FONT_RE.search(font) and not _FONT_RE.search(content):
                content_len = data[off2 + 3]
                content_end = off2 + 4 + content_len * 2
                # 座標は content_end + 17 バイト目
                coord_off = content_end + 17
                x, y = 0.0, 0.0
                if coord_off + 16 <= len(data):
                    try:
                        x = struct.unpack_from("<d", data, coord_off)[0]
                        y = struct.unpack_from("<d", data, coord_off + 8)[0]
                        if not (-1e6 < x < 1e6 and -1e6 < y < 1e6):
                            x, y = 0.0, 0.0
                    except struct.error:
                        pass
                records.append((x, y, off2, content))
                j += 2
            else:
                j += 1

        if not records:
            return []

        # ── 印刷範囲フィルタ（二重枠外のテキストを除外）──
        # ケーブル・配管パターンに一致するテキストを「シード」として
        # 印刷範囲の中心を推定し、そこから遠すぎるテキストを除外する。
        # （シンボルライブラリ等、図面領域外のレイヤテキストを排除）
        seed_xs = [r[0] for r in records
                   if (CONDUIT_PATTERN.search(r[3]) or CABLE_PATTERN.search(r[3])
                       or LOCATION_PATTERN.search(r[3]))
                   and r[0] != 0.0]
        seed_ys = [r[1] for r in records
                   if (CONDUIT_PATTERN.search(r[3]) or CABLE_PATTERN.search(r[3])
                       or LOCATION_PATTERN.search(r[3]))
                   and r[1] != 0.0]
        if seed_xs and seed_ys:
            x_lo = min(seed_xs) - 200.0
            x_hi = max(seed_xs) + 200.0
            y_lo = min(seed_ys) - 200.0
            y_hi = max(seed_ys) + 200.0
            records = [
                r for r in records
                if (r[0] == 0.0 or x_lo <= r[0] <= x_hi)
                and (r[1] == 0.0 or y_lo <= r[1] <= y_hi)
            ]

        # ── 配管→ケーブルの順序修正 ──
        # ファイル上ではケーブルが配管より先に現れる場合があるため順序を修正する。
        def _is_conduit(t: str) -> bool:
            return bool(CONDUIT_PATTERN.search(t))

        def _is_cable(t: str) -> bool:
            return bool(CABLE_PATTERN.search(t))

        _COUNT_RE = re.compile(r'(\d+)[×x×]\d+\s*\(')
        _FEP_RE   = re.compile(r'FEP')

        def _cable_count(t: str) -> int:
            m = _COUNT_RE.search(t)
            return int(m.group(1)) if m else 99

        def _fep_adjacent_cable_types(recs_list: list, fep_idx: int) -> set:
            """FEP配管の直後に隣接するケーブル種別セット（次の配管または設置場所まで）"""
            types: set = set()
            for j in range(fep_idx + 1, min(fep_idx + 8, len(recs_list))):
                jx, jy, joff, jt = recs_list[j]
                if _is_conduit(jt) or LOCATION_PATTERN.search(jt):
                    break
                if _is_cable(jt):
                    m = CABLE_PATTERN.search(jt)
                    if m:
                        types.add(normalize_cable(m.group(1)))
            return types

        # ── Phase 1: 近傍配管スワップ（X距離 < 30mm）──
        records.sort(key=lambda r: r[2])
        recs = list(records)
        i = 0
        while i < len(recs) - 1:
            rx, ry, roff, rt = recs[i]
            if _is_cable(rt) and not _is_conduit(rt):
                for k in range(i + 1, min(i + 6, len(recs))):
                    kx, ky, koff, kt = recs[k]
                    if _is_conduit(kt) and abs(kx - rx) < 30:
                        cable_rec = recs.pop(i)
                        recs.insert(k, cable_rec)
                        i = k + 1
                        break
                else:
                    i += 1
            else:
                i += 1

        # ── Phase 2: 1×1ケーブルをFEP配管へ補完割り当て ──
        # FEP配管の直後に隣接していない count=1 ケーブルがあり、
        # そのFEP配管がその種別のケーブルをまだ持っていない場合に移動する。
        # （JWW図面でケーブルラベルがルート列に配置される描画慣習への対応）
        changed = True
        while changed:
            changed = False
            fep_idxs = [idx for idx, (_, _, _, t) in enumerate(recs)
                        if _is_conduit(t) and _FEP_RE.search(t)]
            if not fep_idxs:
                break

            for i, (rx, ry, roff, rt) in enumerate(recs):
                if not (_is_cable(rt) and not _is_conduit(rt)):
                    continue
                if _cable_count(rt) != 1:
                    continue
                ct_m = CABLE_PATTERN.search(rt)
                if not ct_m:
                    continue
                cable_norm = normalize_cable(ct_m.group(1))

                # このケーブルより後ろにあるFEP配管を探す
                feps_after = [fp for fp in fep_idxs if fp > i]
                if not feps_after:
                    continue

                # FEPとのX距離が全て30mm以上のときのみ対象
                if any(abs(recs[fp][0] - rx) < 30 for fp in feps_after):
                    continue

                # 受け入れ先FEP: この種別をまだ持っていない最初のFEP
                target_fep = None
                for fp in sorted(feps_after):
                    if cable_norm not in _fep_adjacent_cable_types(recs, fp):
                        target_fep = fp
                        break

                if target_fep is not None:
                    cable_rec = recs.pop(i)
                    adj = target_fep - (1 if i < target_fep else 0)
                    recs.insert(adj + 1, cable_rec)
                    changed = True
                    break  # recs 変更後はリストを再スキャン

        seen: set[tuple] = set()
        lines: list[str] = []
        for rx, ry, _, text in recs:
            text = text.replace("　", " ").replace(" ", " ").strip()
            if len(text) < 2:
                continue
            dedup_key = (text, round(rx, -1), round(ry, -1))
            if dedup_key in seen:
                continue
            seen.add(dedup_key)
            lines.append(text)

        return lines

    except Exception as e:
        print(f"JWW読込エラー: {e}", file=sys.stderr)
        return []


# ─────────────────────────────────────────
#  画像OCRからテキスト抽出
# ─────────────────────────────────────────

def extract_text_from_image(filepath: str) -> list[str]:
    """pytesseract + OpenCVで画像からテキスト抽出"""
    try:
        import cv2
        import numpy as np
        import pytesseract
    except ImportError:
        print("画像OCRには pytesseract と opencv-python が必要です:")
        print("  pip install pytesseract opencv-python")
        print("  Tesseract本体: https://github.com/UB-Mannheim/tesseract/wiki")
        return []

    img = cv2.imread(filepath)
    if img is None:
        print(f"画像を開けません: {filepath}")
        return []

    # 緑色テキストのみ抽出 (HSV色空間でマスク)
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    # 緑系: H=40-80, 青緑系: H=80-100
    mask1 = cv2.inRange(hsv, (35, 40, 40), (85, 255, 255))
    # シアン系 (JwCADの緑表示)
    mask2 = cv2.inRange(hsv, (85, 40, 40), (100, 255, 255))
    mask = cv2.bitwise_or(mask1, mask2)

    # マスク適用
    green_only = cv2.bitwise_and(img, img, mask=mask)

    # グレースケール→二値化
    gray = cv2.cvtColor(green_only, cv2.COLOR_BGR2GRAY)
    _, binary = cv2.threshold(gray, 30, 255, cv2.THRESH_BINARY)

    # OCR (日本語)
    try:
        text = pytesseract.image_to_string(
            binary,
            lang="jpn+eng",
            config="--psm 6"
        )
    except Exception:
        # 日本語モデルなしの場合は英語のみ
        text = pytesseract.image_to_string(binary, config="--psm 6")

    lines = [l.strip() for l in text.splitlines() if l.strip()]
    return lines


# ─────────────────────────────────────────
#  PDFからテキスト抽出
# ─────────────────────────────────────────

def extract_text_from_pdf(filepath: str) -> list[str]:
    """
    pdfplumberで単語座標を使いPDFテキストを抽出する。
    JwCAD図面は複数の注記が同じ高さに横並びになっているため、
    単語のX座標のクラスタリングで列を分離して別行として扱う。
    """
    try:
        import pdfplumber
    except ImportError:
        print("PDFには pdfplumber が必要です: pip install pdfplumber")
        return []

    Y_TOL = 4    # 同一行とみなすY座標の許容誤差(pt) ← extract_words に渡す
    NO_SP = 2    # この距離未満ならスペースなしで結合（文字間）
    COL_SPLIT = 55  # この距離以上なら別列に分割(pt)
    Y_ROW = 10   # 列境界検出用：この距離以内のY差なら同一行とみなす(pt)

    def _is_black(color) -> bool:
        """テキスト色が黒に近いか判定（黒文字はカウント対象外）"""
        if color is None:
            return False
        if isinstance(color, (int, float)):
            return float(color) < 0.1
        if isinstance(color, (list, tuple)):
            c = [float(x) for x in color]
            if len(c) == 1:
                return c[0] < 0.1
            if len(c) == 3:   # RGB
                return max(c) < 0.1
            if len(c) == 4:   # CMYK
                return c[3] > 0.9 and max(c[:3]) < 0.1
        return False

    all_lines = []

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            # 黒文字を除外してから単語抽出
            filtered = page.filter(
                lambda obj: obj["object_type"] != "char"
                or not _is_black(obj.get("non_stroking_color"))
            )
            words = filtered.extract_words(
                x_tolerance=3,
                y_tolerance=Y_TOL,
                keep_blank_chars=False,
            )
            if not words:
                continue

            # Y座標でグループ化
            rows: dict[float, list[dict]] = {}
            for w in words:
                y_key = round(w["top"] / Y_TOL) * Y_TOL
                rows.setdefault(y_key, []).append(w)

            # 各Y行を列チャンク (col_x, y, text) に分割して収集
            chunks: list[tuple[float, float, str]] = []
            for y in sorted(rows):
                row_words = sorted(rows[y], key=lambda w: w["x0"])

                col_x_start = row_words[0]["x0"]
                cur = ""
                prev_x1 = None

                for w in row_words:
                    if prev_x1 is None:
                        cur = w["text"]
                    else:
                        gap = w["x0"] - prev_x1
                        if gap >= COL_SPLIT:
                            # 大きなギャップ → 列の切れ目
                            if cur.strip():
                                chunks.append((col_x_start, y, cur.strip()))
                            cur = w["text"]
                            col_x_start = w["x0"]
                        elif gap < NO_SP:
                            cur += w["text"]
                        else:
                            cur += " " + w["text"]
                    prev_x1 = w["x1"]

                if cur.strip():
                    chunks.append((col_x_start, y, cur.strip()))

            # 列ごとに上から下へ処理する
            # Y_ROW pt以内なら「同一行」とみなして列境界を検出し、
            # 各列を上→下の順に並べることで配管→内線の対応を正しく保つ
            if chunks:
                # ① Y でソートして Y_ROW 許容でグループ化
                by_y = sorted(chunks, key=lambda c: c[1])
                row_groups: list[list] = []
                cur_grp = [by_y[0]]
                for c in by_y[1:]:
                    if c[1] - cur_grp[-1][1] <= Y_ROW:
                        cur_grp.append(c)
                    else:
                        row_groups.append(cur_grp)
                        cur_grp = [c]
                row_groups.append(cur_grp)

                # ② 各グループ内でX方向に COL_SPLIT 以上のギャップを列境界とする
                split_mids: list[float] = []
                for grp in row_groups:
                    sx = sorted(grp, key=lambda c: c[0])
                    for i in range(len(sx) - 1):
                        gap = sx[i + 1][0] - sx[i][0]
                        if gap >= COL_SPLIT:
                            split_mids.append((sx[i][0] + sx[i + 1][0]) / 2)

                # ③ 中点をクラスタリングして列境界リストを確定
                boundaries: list[float] = []
                if split_mids:
                    split_mids.sort()
                    cluster = [split_mids[0]]
                    for mid in split_mids[1:]:
                        if mid - cluster[-1] < COL_SPLIT:
                            cluster.append(mid)
                        else:
                            boundaries.append(sum(cluster) / len(cluster))
                            cluster = [mid]
                    boundaries.append(sum(cluster) / len(cluster))

                def col_idx(x: float) -> int:
                    return sum(1 for b in boundaries if x > b)

                chunks_sorted = sorted(chunks, key=lambda c: (col_idx(c[0]), c[1]))
                all_lines.extend(c[2] for c in chunks_sorted)

    return [_clean_pdf_line(l) for l in all_lines]


def _clean_pdf_line(line: str) -> str:
    """
    PDF抽出時に文字間に入った余分なスペースを除去する。
    ケーブル・配管・数量の表記パターン内のスペースをつぶす。
    """
    # IV/CVT ケーブル種: 数字・小数点・スペースが混入しても除去
    # 例: IV1 4sq → IV14sq, IV 5 .5 s q → IV5.5sq
    def collapse(m):
        return re.sub(r'\s+', '', m.group(0))
    line = re.sub(r'(?:IV|CVT|CVD|CV)\s*[\d\s.]+\s*s\s*q', collapse, line, flags=re.IGNORECASE)
    # 6.6kV: 6 . 6 kV → 6.6kV
    line = re.sub(r'6\s*\.\s*6\s*k\s*[Vv]', '6.6kV', line)
    # FEP/HIVE括弧: FEP ( 80 ) → FEP(80)
    line = re.sub(r'(FEP|HIVE)\s*\(\s*(\d+)\s*\)', r'\1(\2)', line, flags=re.IGNORECASE)
    # 数量: 9 × 1 ( 9 m ) → 9×1(9m)
    line = re.sub(r'(\d+)\s*[×x×]\s*(\d+)\s*\(\s*(\d+)\s*m\s*\)', r'\1×\2(\3m)', line)
    return line


# ─────────────────────────────────────────
#  テキストファイル読込
# ─────────────────────────────────────────

def extract_text_from_txt(filepath: str) -> list[str]:
    """テキストファイルから行ごとに読込（手動入力用）"""
    try:
        with open(filepath, encoding="utf-8-sig") as f:
            return [l.strip() for l in f if l.strip()]
    except UnicodeDecodeError:
        with open(filepath, encoding="shift_jis") as f:
            return [l.strip() for l in f if l.strip()]


# ─────────────────────────────────────────
#  テキストパターン解析
# ─────────────────────────────────────────

# 配管パターン例:
#   露出立下 HIVE(70) 8×1(8m)
#   埋設 FEP(80) 10×1(10m)
#   露出 FEP(80) 3×1(3m)
# ×の前が本数、()内が長さ

CONDUIT_PATTERN = re.compile(
    r"(露出|埋設|隠蔽)[立下上配管蔽]*\s*"
    r"("
      # 長形式プレフィックス付き or タイプコードのみ (サイズ付き)
      r"(?:(?:ねじなし電線管|厚鋼電線管|金属製可とう電線管)\s*)?"
      r"(?:HIVE|VE|FEP|PFD|PF|PV|PZ|KMS|KMV)\s*\(?\s*\d+\s*\)?"
      r"|(?:(?:ねじなし電線管|厚鋼電線管)\s*)?[GCE]\s*\(?\s*\d+\s*\)?"  # G/C/E 鋼管
      r"|フレキ(?:\s*\(?\s*\d+\s*\)?)?"                    # フレキ → KMS
      r"|金属製可とう電線管(?:\s*\(?\s*\d+\s*\)?)?"        # 金属製可とう電線管 → KMS
      r"|厚鋼|G管|薄鋼|C管|E管|防水プリカ"                # 名称のみ (サイズなし)
    r")"
    r"\s*\d+\s*[×x×]\s*\d+\s*\((\d+)m\)",
    re.IGNORECASE
)

# ケーブルパターン例:
#   6.6kV CVT38sq 9×1(9m)
#   CVT150sq 3×1(3m)
#   IV14sq 2×1(2m)
#   IV5.5sq 2×1(2m)
#   IV38sq 2×1(2m)

CABLE_PATTERN = re.compile(
    r"((?:6\.6kV\s*)?"                    # 6.6kV プレフィックス（任意）
    r"(?:CVT|CVD|CV|DV|IV)\s*[\d.]+\s*"   # ケーブル種 + サイズ
    r"(?:sq)?\s*(?:-\s*\d+C)?)"          # sq サフィックス（任意）・-3C/-2C（任意）
    r"\s*(\d+)\s*[×x×]\s*\d+\s*\((\d+)m\)",
    re.IGNORECASE
)

# 配管なし設置場所パターン例:
#   新設コンクリートポール内立下 IV14sq 9×1(9m)  ← 同行にケーブル
#   受麦電設備内                                   ← 次行以降にケーブル
#   充電器内配線
#   キュービクル内配線 / QB内配線
LOCATION_PATTERN = re.compile(
    r"(?:新設|既設)?"
    r"(?:[^\s]*)"                         # 「新設コンクリートポール」など任意の前置詞
    r"(?:"
      r"ポール内|電柱内|柱内"             # 電柱・ポール系（総称）
      r"|CP内"                            # コンクリートポール（CP柱・コンクリートポール）
      r"|NAポール内|鋼管柱内|鋼管ポール内"  # NAポール・鋼管ポール
      r"|SH-?\d+内"                      # SH-6・SH-7などの鋼管柱型番
      r"|充電器内|EV充電器内"             # 充電器系
      r"|キュービクル内|QB内|QC内|キュビクル内"  # キュービクル系
      r"|受[変麦]電設備内|受電設備内|変電設備内"  # 受変電設備系
      r"|制御盤内|分電盤内|端子盤内|盤内"  # 盤類
    r")"
    r"[配線立下]*"                         # 「配線」「立下」などの後置詞
)

# バリカーパターン例:
#   バリカー ×3
#   バリカー 3台
#   バリカー 3基
#   バリカー(電動) 2基
#   電動バリカー ×2
#   バリカー設置 ×1
#   新設バリカー 2本
BOLLARD_PATTERN = re.compile(
    r"(?:新設|既設|撤去)?"                           # 任意のプレフィックス
    r"(?:電動|固定|手動)?"                           # 任意の種別修飾
    r"バリカー"                                      # キーワード
    r"(?:[（(][^）)]*[）)])?"                        # 任意の括弧内情報: (電動)など
    r"(?:設置|撤去|新設)?"                           # 任意の動詞
    r"\s*"
    r"(?:[×x×＊*]?\s*(\d+)|(\d+)\s*(?:台|基|本|個|柱|箇所|ヶ所|か所))?",  # 数量
    re.IGNORECASE
)

def normalize_conduit(raw_type: str) -> str:
    """配管種を正規化。VE→HIVE、長形式プレフィックス除去、フレキ→KMS など。"""
    s = re.sub(r"\s+", "", raw_type.strip())
    # 長形式プレフィックスを除去してからマッチ
    s = re.sub(r"^ねじなし電線管", "", s)
    s = re.sub(r"^厚鋼電線管", "", s)
    s = re.sub(r"^金属製可とう電線管", "", s)
    # フレキ → KMS
    m = re.match(r"フレキ\(?(\d*)\)?", s)
    if m:
        return f"KMS({m.group(1)})" if m.group(1) else "KMS"
    # VE / HIVE → HIVE
    m = re.match(r"(?:VE|HIVE)\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"HIVE({m.group(1)})"
    # FEP
    m = re.match(r"FEP\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"FEP({m.group(1)})"
    # PFD (PFより先にマッチ)
    m = re.match(r"PFD\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"PFD({m.group(1)})"
    # PF
    m = re.match(r"PF\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"PF({m.group(1)})"
    # PV (防水プリカ)
    m = re.match(r"PV\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"PV({m.group(1)})"
    # PZ (プリカ)
    m = re.match(r"PZ\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"PZ({m.group(1)})"
    # KMS (金属製可とう電線管・フレキ)
    m = re.match(r"KMS\(?(\d*)\)?", s, re.IGNORECASE)
    if m:
        return f"KMS({m.group(1)})" if m.group(1) else "KMS"
    # KMV
    m = re.match(r"KMV\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"KMV({m.group(1)})"
    # G管 / 厚鋼 / 厚鋼電線管 / G(54)
    if re.match(r"厚鋼|G管", s):
        return "G管"
    m = re.match(r"G\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"G({m.group(1)})"
    # C管 / 薄鋼 / C(25)
    if re.match(r"薄鋼|C管", s):
        return "C管"
    m = re.match(r"C\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"C({m.group(1)})"
    # E管 / ねじなし電線管 / E(25)
    if re.match(r"E管", s):
        return "E管"
    m = re.match(r"E\(?(\d+)\)?", s, re.IGNORECASE)
    if m:
        return f"E({m.group(1)})"
    # 防水プリカ
    if re.match(r"防水プリカ", s):
        return "防水プリカ"
    return s

def normalize_cable(raw_type: str) -> str:
    """ケーブル種を正規化。IV14 → IV14sq など sq なしを補完。"""
    s = re.sub(r"\s+", "", raw_type.strip())
    # IV14 / IV5.5 など sq なしを補完 (-NC サフィックスなし・sq なし の場合のみ)
    s = re.sub(r"^(IV[\d.]+)$", r"\1sq", s, flags=re.IGNORECASE)
    return s

def parse_bollard(line: str) -> Optional[tuple[str, int]]:
    """
    1行からバリカー情報を抽出する。
    戻り値: (ラベル文字列, 数量) または None
    """
    bm = BOLLARD_PATTERN.search(line)
    if not bm:
        return None

    # マッチした文字列が "バリカー" を含むか確認（空マッチ対策）
    matched_str = bm.group(0)
    if "バリカー" not in matched_str:
        return None

    # 数量取得
    qty_str = bm.group(1) or bm.group(2)
    qty = int(qty_str) if qty_str else 1

    # ラベル: 行全体からバリカー周辺を取り出す（括弧種別含む）
    label_match = re.search(
        r"(?:新設|既設|撤去)?(?:電動|固定|手動)?バリカー(?:[（(][^）)]*[）)])?(?:設置|撤去|新設)?",
        line
    )
    label = label_match.group(0).strip() if label_match else "バリカー"

    return label, qty


def parse_location(line: str) -> Optional[str]:
    """設置場所ヘッダー行を検出してラベルを返す"""
    m = LOCATION_PATTERN.search(line)
    if not m:
        return None
    # マッチ箇所を含む意味のある部分を返す（長すぎる場合は末尾20文字に収める）
    label = m.group(0).strip()
    return label if label else None


def parse_lines(lines: list[str]) -> tuple[list[ConduitEntry], list[CableEntry], list[FreeCableEntry], list[tuple[str, int]]]:
    """
    テキスト行リストから配管・ケーブル・配管なしケーブル・バリカーを抽出する。
    ※「双方に記載」マーカーが付いたブロックはケーブル種+長さの組み合わせで重複排除する。
    """
    conduits: list[ConduitEntry] = []
    cables: list[CableEntry] = []
    free_cables: list[FreeCableEntry] = []
    bollards: list[tuple[str, int]] = []

    block_conduit: Optional[ConduitEntry] = None
    block_location: Optional[str] = None

    # 「双方に記載」重複排除用
    block_cable_start: int = 0        # 現ブロック開始時点の cables[] インデックス
    block_conduit_idx: Optional[int] = None  # 現ブロックの conduit の conduits[] インデックス
    block_free_start: int = 0         # 現ブロック開始時点の free_cables[] インデックス
    dual_cable_sigs: set = set()      # 既に計上済みの「双方に記載」ブロックのケーブル署名

    def reset_block():
        nonlocal block_conduit, block_location, block_cable_start, block_conduit_idx, block_free_start
        block_conduit = None
        block_location = None
        block_conduit_idx = None
        block_cable_start = len(cables)
        block_free_start = len(free_cables)

    for line in lines:
        line = line.replace("　", " ")  # 全角スペース→半角
        line = line.translate(str.maketrans(
            "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
            "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ"
            "０１２３４５６７８９－（）",
            "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
            "abcdefghijklmnopqrstuvwxyz"
            "0123456789-()"
        ))

        if re.search(r"既設|撤去", line):
            continue

        # ※双方に記載 → ケーブル署名で重複チェックし、重複なら直前ブロックを除去
        if re.search(r"双方に記載", line):
            sig = frozenset((c.cable_type, c.length) for c in cables[block_cable_start:])
            if sig and sig in dual_cable_sigs:
                # 重複 → 今ブロックの配管・ケーブル・配管なしケーブルをすべて取り消す
                if block_conduit_idx is not None:
                    del conduits[block_conduit_idx:]
                del cables[block_cable_start:]
                del free_cables[block_free_start:]
            else:
                if sig:
                    dual_cable_sigs.add(sig)
            reset_block()
            continue

        # バリカーマッチ
        bollard = parse_bollard(line)
        if bollard:
            bollards.append(bollard)
            reset_block()
            continue

        # 架空ケーブル（配管に入らない空中配線）→ 総量のみ計上してブロックリセット
        if re.match(r"架空", line):
            for km in CABLE_PATTERN.finditer(line):
                cable_type = normalize_cable(km.group(1))
                cable_len = int(km.group(3))
                cables.append(CableEntry(cable_type, cable_len))
            reset_block()
            continue

        # 設置場所ヘッダー検出（配管より先にチェック）
        location = parse_location(line)
        if location:
            block_conduit = None
            block_conduit_idx = None
            block_location = location
            block_cable_start = len(cables)
            block_free_start = len(free_cables)
            for km in CABLE_PATTERN.finditer(line):
                cable_type = normalize_cable(km.group(1))
                cable_len = int(km.group(3))
                cables.append(CableEntry(cable_type, cable_len))
                free_cables.append(FreeCableEntry(block_location, cable_type, cable_len))
            continue

        # 配管マッチ
        cm = CONDUIT_PATTERN.search(line)
        if cm:
            install = cm.group(1)
            conduit_raw = cm.group(2)
            length = int(cm.group(3))
            conduit_type = normalize_conduit(conduit_raw)
            entry = ConduitEntry(install, conduit_type, length)
            conduits.append(entry)
            block_conduit = entry
            block_conduit_idx = len(conduits) - 1
            block_location = None
            block_cable_start = len(cables)
            block_free_start = len(free_cables)
            for km in CABLE_PATTERN.finditer(line):
                cable_type = normalize_cable(km.group(1))
                cable_len = int(km.group(3))
                cables.append(CableEntry(cable_type, cable_len))
                if cable_type not in entry.cables:
                    entry.cables.append(cable_type)
            continue

        # ケーブルマッチ
        matched_cable = False
        for km in CABLE_PATTERN.finditer(line):
            cable_type = normalize_cable(km.group(1))
            cable_len = int(km.group(3))
            cables.append(CableEntry(cable_type, cable_len))
            if block_conduit:
                if cable_type not in block_conduit.cables:
                    block_conduit.cables.append(cable_type)
            elif block_location:
                free_cables.append(FreeCableEntry(block_location, cable_type, cable_len))
            matched_cable = True

        if not matched_cable:
            if line and not re.match(r"[・※▼▶　\s#]", line):
                reset_block()

    return conduits, cables, free_cables, bollards


# ─────────────────────────────────────────
#  集計
# ─────────────────────────────────────────

def aggregate_free_cables(free_cables: list[FreeCableEntry]) -> list[tuple[str, str, int]]:
    """
    配管なしケーブルを (設置場所, ケーブル種) でまとめて合計長を集計。
    → [(location, cable_type, total_length), ...] 挿入順
    """
    totals: dict[tuple[str, str], int] = defaultdict(int)
    for fc in free_cables:
        totals[(fc.location, fc.cable_type)] += fc.length
    return [(loc, ctype, ln) for (loc, ctype), ln in totals.items()]


def aggregate_bollards(bollards: list[tuple[str, int]]) -> dict[str, int]:
    """バリカーをラベル別に合計数量集計: {ラベル: 合計数量}"""
    totals: dict[str, int] = defaultdict(int)
    for label, qty in bollards:
        totals[label] += qty
    return dict(totals)


def aggregate(
    conduits: list[ConduitEntry],
    cables: list[CableEntry]
) -> tuple[dict, list]:
    """
    配線種集計: {cable_type: total_length}
    配管種集計: 入っているケーブルが同一の区間をまとめて集計
      → [(install_type, conduit_type, sorted_cables_tuple, total_length), ...]
    """
    cable_totals: dict[str, int] = defaultdict(int)
    for c in cables:
        cable_totals[c.cable_type] += c.length

    # (install_type, conduit_type, cables_key) → total_length
    conduit_map: dict[tuple, int] = defaultdict(int)
    for c in conduits:
        cables_key = tuple(sorted(c.cables))
        key = (c.install_type, c.conduit_type, cables_key)
        conduit_map[key] += c.length

    # リスト形式に変換し、配管サイズの数字で降順ソート
    conduit_rows = [
        (inst, ctype, list(cables_key), length)
        for (inst, ctype, cables_key), length in conduit_map.items()
    ]

    def _conduit_size(row):
        m = re.search(r'\d+', row[1])
        return int(m.group()) if m else 0

    conduit_rows.sort(key=_conduit_size, reverse=True)

    return dict(cable_totals), conduit_rows


# ─────────────────────────────────────────
#  Excel出力
# ─────────────────────────────────────────

def export_to_excel(
    cable_totals: dict,
    conduit_totals: dict,
    accessory_info: dict,
    output_path: str
):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    except ImportError:
        print("Excelエクスポートには openpyxl が必要です: pip install openpyxl")
        export_to_csv(cable_totals, conduit_totals, accessory_info, output_path.replace(".xlsx", ".csv"))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "サマリ"

    # スタイル定義
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    def cell(row, col, value, font=None, fill=None, align=None, border_=True):
        c = ws.cell(row=row, column=col, value=value)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        if border_:
            c.border = border
        return c

    row = 1

    # タイトル
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="サマリ")
    c.font = Font(bold=True, size=14)
    c.alignment = center
    row += 1

    # ─ 配線種セクション ─
    ws.merge_cells(f"A{row}:D{row}")
    cell(row, 1, "配線種", bold, header_fill, center)
    row += 1

    # ケーブル順序 (図面の慣例に合わせてソート)
    cable_order = ["6.6kVCVT38sq", "CVT150sq", "IV38sq", "IV14sq", "IV5.5sq"]
    ordered_cables = cable_order + [k for k in cable_totals if k not in cable_order]

    for cable_type in ordered_cables:
        if cable_type not in cable_totals:
            continue
        length = cable_totals[cable_type]
        cell(row, 1, "", border_=True)  # 空白
        cell(row, 2, cable_type, align=center)
        ws.merge_cells(f"C{row}:D{row}")
        cell(row, 3, f"全長{length}m", align=center)
        ws.cell(row=row, column=4).border = border
        row += 1

    row += 1  # 空行

    # ─ 配管種セクション ─
    ws.merge_cells(f"A{row}:D{row}")
    cell(row, 1, "配管種", bold, header_fill, center)
    row += 1

    # 設置種別ごとにグループ化して出力
    for install_type in ["露出", "埋設"]:
        entries = [(inst, ctype, cab, ln) for (inst, ctype, cab, ln) in conduit_totals if inst == install_type]
        for inst, conduit_type, cables_in_conduit, length in entries:
            cable_note = " / ".join(cables_in_conduit) if cables_in_conduit else ""
            cell(row, 1, install_type, align=center)
            cell(row, 2, conduit_type, align=center)
            cell(row, 3, f"全長{length}m", align=center)
            cell(row, 4, f"({cable_note})" if cable_note else "", align=left)
            row += 1

    row += 1

    # ─ 付帯設備セクション ─
    ws.merge_cells(f"A{row}:D{row}")
    cell(row, 1, "付帯設備", bold, header_fill, center)
    row += 1

    if accessory_info:
        for key, value in accessory_info.items():
            cell(row, 1, key, align=center)
            ws.merge_cells(f"B{row}:D{row}")
            cell(row, 2, value, align=center)
            ws.cell(row=row, column=3).border = border
            ws.cell(row=row, column=4).border = border
            row += 1
    else:
        ws.merge_cells(f"A{row}:D{row}")
        cell(row, 1, "（検出されませんでした）", align=left)
        row += 1

    # 列幅調整
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30

    wb.save(output_path)
    print(f"Excel出力完了: {output_path}")


def export_to_csv(cable_totals, conduit_totals, accessory_info, output_path):
    import csv
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["サマリ"])
        w.writerow(["配線種"])
        for cable_type, length in cable_totals.items():
            w.writerow(["", cable_type, f"全長{length}m"])
        w.writerow([])
        w.writerow(["配管種"])
        for inst, conduit, cables, length in conduit_totals:
            cables_note = "/".join(cables)
            w.writerow([inst, conduit, f"全長{length}m", cables_note])
        w.writerow([])
        w.writerow(["付帯設備"])
        for k, v in (accessory_info or {}).items():
            w.writerow([k, v])
    print(f"CSV出力完了: {output_path}")


# ─────────────────────────────────────────
#  メイン
# ─────────────────────────────────────────

SUPPORTED_EXT = {".jww", ".jw_", ".txt", ".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}

def process_file(filepath: str) -> tuple[list[str], str]:
    """ファイルの種類に応じてテキストを抽出"""
    ext = Path(filepath).suffix.lower()
    if ext in (".jww", ".jw_"):
        print(f"JWWファイルを解析中: {filepath}")
        lines = extract_text_from_jww(filepath)
    elif ext == ".txt":
        print(f"テキストファイルを読込中: {filepath}")
        lines = extract_text_from_txt(filepath)
    elif ext == ".pdf":
        print(f"PDFを解析中: {filepath}")
        lines = extract_text_from_pdf(filepath)
    elif ext in (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"):
        print(f"画像をOCR処理中: {filepath}")
        lines = extract_text_from_image(filepath)
    else:
        print(f"未対応のファイル形式: {ext}")
        lines = []
    return lines, ext


def main():
    if len(sys.argv) < 2:
        print("使い方: python jwcad_summary.py <ファイルパス> [出力ファイル.xlsx]")
        print("対応形式: .jww  .txt  .png  .jpg  .bmp  .tif")
        print()
        print("例:")
        print("  python jwcad_summary.py 図面.jww")
        print("  python jwcad_summary.py 図面.jww 集計表.xlsx")
        print("  python jwcad_summary.py 図面テキスト.txt 集計表.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]
    if not os.path.exists(input_path):
        print(f"ファイルが見つかりません: {input_path}")
        sys.exit(1)

    output_path = sys.argv[2] if len(sys.argv) >= 3 else \
        str(Path(input_path).with_suffix(".xlsx"))

    # テキスト抽出
    lines, ext = process_file(input_path)

    if not lines:
        print("テキストを抽出できませんでした。")
        sys.exit(1)

    print(f"\n抽出テキスト ({len(lines)}行):")
    for l in lines[:30]:
        print(f"  {l}")
    if len(lines) > 30:
        print(f"  ... 他{len(lines)-30}行")

    # 解析
    conduits, cables, free_cables, bollards = parse_lines(lines)

    print(f"\n配管エントリ: {len(conduits)}件")
    for c in conduits:
        print(f"  [{c.install_type}] {c.conduit_type} {c.length}m  ← {c.cables}")

    print(f"\nケーブルエントリ: {len(cables)}件")
    for c in cables:
        print(f"  {c.cable_type} {c.length}m")

    print(f"\n配管なしケーブル: {len(free_cables)}件")
    for fc in free_cables:
        print(f"  [{fc.location}] {fc.cable_type} {fc.length}m")

    print(f"\nバリカーエントリ: {len(bollards)}件")
    for label, qty in bollards:
        print(f"  {label}: {qty}台")

    # 集計
    cable_totals, conduit_totals = aggregate(conduits, cables)
    free_cable_totals = aggregate_free_cables(free_cables)
    bollard_totals = aggregate_bollards(bollards)

    print("\n─── 集計結果 ───")
    print("配線種:")
    for k, v in cable_totals.items():
        print(f"  {k}: 全長{v}m")
    print("配管種:")
    for inst, ctype, cabs, length in conduit_totals:
        print(f"  {inst} {ctype}: 全長{length}m  ← {cabs}")
    print("配管なしケーブル:")
    for loc, ctype, length in free_cable_totals:
        print(f"  [{loc}] {ctype}: 全長{length}m")
    print("付帯設備（バリカー）:")
    for label, qty in bollard_totals.items():
        print(f"  {label}: {qty}台")

    # 付帯設備: バリカー集計結果を渡す
    accessory_info = {label: f"{qty}台" for label, qty in bollard_totals.items()}

    # 出力
    if output_path.endswith(".xlsx"):
        export_to_excel(cable_totals, conduit_totals, accessory_info, output_path)
    else:
        export_to_csv(cable_totals, conduit_totals, accessory_info, output_path)


if __name__ == "__main__":
    main()
