"""
JwCAD 配線・配管集計アプリ (Streamlit)
"""

import io
import json
import tempfile
import os
from pathlib import Path
import streamlit as st
import pandas as pd

from jwcad_summary import (
    extract_text_from_jww,
    extract_text_from_txt,
    extract_text_from_pdf,
    parse_lines,
    aggregate,
    aggregate_free_cables,
    aggregate_bollards,
    PanelEntry,
)

# ─────────────────────────────────────────
#  盤寸法キャッシュ
# ─────────────────────────────────────────

_CACHE_PATH = Path(__file__).parent / "panel_dimensions_cache.json"

# 既知モデルの初期データ
# dims: タテ×ヨコ×深さ mm / body: 本体材質 / plate: 中板材質
_BUILTIN_CACHE: dict[str, dict] = {
    # 日東工業 OMS-B 引込計器盤キャビネット（鉄製本体・木製中板）
    "OMS-121B": {"dims": "1200×400×200", "body": "鋼板（鉄製）", "plate": "木製"},
    "OMS-21B":  {"dims": "800×500×200",  "body": "鋼板（鉄製）", "plate": "木製"},
    "OMS-11B":  {"dims": "800×400×200",  "body": "鋼板（鉄製）", "plate": "木製"},
    "OMS-12B":  {"dims": "1000×400×200", "body": "鋼板（鉄製）", "plate": "木製"},
    "OMS-251B": {"dims": "1000×500×200", "body": "鋼板（鉄製）", "plate": "木製"},
    # 日東工業 OPK-A キー付耐候プラボックス（屋根付）
    "OPK18-35A": {"dims": "500×300×180", "body": "AAS樹脂",     "plate": "木製"},
}

def _load_cache() -> dict[str, dict]:
    cache: dict[str, dict] = {k: dict(v) for k, v in _BUILTIN_CACHE.items()}
    if _CACHE_PATH.exists():
        try:
            stored = json.loads(_CACHE_PATH.read_text(encoding="utf-8"))
            for k, v in stored.items():
                # 旧フォーマット（文字列）との互換性
                if isinstance(v, str):
                    cache[k] = {"dims": v, "body": "", "plate": ""}
                else:
                    cache[k] = v
        except Exception:
            pass
    return cache

def _save_cache(cache: dict[str, dict]) -> None:
    try:
        diff = {k: v for k, v in cache.items()
                if k not in _BUILTIN_CACHE or _BUILTIN_CACHE[k] != v}
        _CACHE_PATH.write_text(json.dumps(diff, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass

def _lookup_panel_info_web(model: str) -> dict | None:
    """DuckDuckGo 経由で型式の寸法を検索（タテ×ヨコ×深さ）。材質は要手動登録。"""
    try:
        import requests, re as _re
        r = requests.get(
            "https://api.duckduckgo.com/",
            params={"q": f"{model} 盤 寸法 mm タテ ヨコ", "format": "json", "no_html": "1"},
            timeout=6,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        text = r.text
        # タテ×ヨコ×深さ / H×W×D などのパターンを探す
        m = _re.search(r'(?:タテ|縦|[Hh])\s*[=:：]?\s*(\d{3,4})\D+(?:ヨコ|横|[Ww])\s*[=:：]?\s*(\d{3,4})\D+(?:深|フカサ|[Dd])\s*[=:：]?\s*(\d{2,3})', text)
        if not m:
            m = _re.search(r'(\d{3,4})\s*[×xX]\s*(\d{3,4})\s*[×xX]\s*(\d{2,3})', text)
        if m:
            return {"dims": f"{m.group(1)}×{m.group(2)}×{m.group(3)}", "body": "", "plate": ""}
    except Exception:
        pass
    return None

def get_panel_info(model: str, cache: dict[str, dict]) -> dict | None:
    """キャッシュを参照し、なければ Web 検索してパネル情報dictを返す"""
    if model in cache:
        return cache[model]
    info = _lookup_panel_info_web(model)
    if info:
        cache[model] = info
        _save_cache(cache)
    return info

def format_panel_detail(model: str | None, info: dict | None) -> str:
    """付帯設備テーブルの「詳細」列テキストを組み立てる"""
    if not model:
        return "―"
    parts = [model]
    if info:
        if info.get("dims"):
            parts.append(f"{info['dims']}mm")
        if info.get("body"):
            parts.append(f"本体:{info['body']}")
        if info.get("plate"):
            parts.append(f"中板:{info['plate']}")
    return " / ".join(parts)

# ─────────────────────────────────────────
#  ページ設定
# ─────────────────────────────────────────

VERSION = "v2.4"

st.set_page_config(
    page_title="JwCAD 配線・配管 集計ツール",
    page_icon="⚡",
    layout="wide",
)

st.title("⚡ JwCAD 配線・配管 集計ツール")
st.caption(f"JWWファイルまたはテキストファイルを読み込み、配線種・配管種・付帯設備の集計表を作成します。　{VERSION}")

# ─────────────────────────────────────────
#  サイドバー: 入力
# ─────────────────────────────────────────

with st.sidebar:
    st.header("📂 ファイル入力")
    uploaded = st.file_uploader(
        "ファイルをアップロード",
        type=["jww", "jw_", "txt", "pdf"],
        help="JWW・テキスト・PDFに対応しています",
    )

    st.divider()
    st.header("✏️ テキスト直接入力")
    manual_text = st.text_area(
        "図面の文字情報を貼り付け",
        height=300,
        placeholder="例:\n露出立下 HIVE(70) 8×1(8m)\n6.6kV CVT38sq 9×1(9m)\nバリカー(電動) ×3",
    )
    run_manual = st.button("直接入力で集計", type="primary", use_container_width=True)

    st.divider()
    st.info(
        "**対応ファイル**\n"
        "- .jww / .jw_（JwCAD図面）\n"
        "- .pdf\n"
        "- .txt（テキスト）\n\n"
        "**対応表記**\n"
        "- 配管: 露出/埋設 + HIVE・FEP\n"
        "- ケーブル: CVT・IV 系\n"
        "- 付帯設備: バリカー\n"
        "- **「既設」と書かれた行は除外**"
    )

# ─────────────────────────────────────────
#  処理
# ─────────────────────────────────────────

lines = []
source_name = ""

if uploaded is not None:
    ext = Path(uploaded.name).suffix.lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(uploaded.read())
        tmp_path = tmp.name
    try:
        if ext in (".jww", ".jw_"):
            lines = extract_text_from_jww(tmp_path)
        elif ext == ".pdf":
            lines = extract_text_from_pdf(tmp_path)
        else:
            lines = extract_text_from_txt(tmp_path)
    finally:
        os.unlink(tmp_path)
    source_name = uploaded.name

elif run_manual and manual_text.strip():
    lines = [l.strip() for l in manual_text.splitlines() if l.strip()]
    source_name = "直接入力"

# ─────────────────────────────────────────
#  結果表示
# st.empty()で包むことでトップレベルのコンポーネントツリーを固定し
# ファイル切り替え時のReact DOM不整合エラーを防ぐ
# ─────────────────────────────────────────

result_area = st.empty()

if not lines:
    result_area.info("左のサイドバーからファイルをアップロードするか、テキストを貼り付けて「集計」ボタンを押してください。")
else:
    with result_area.container():
        # 解析
        conduits, cables, free_cables, bollards, panels = parse_lines(lines)
        cable_totals, conduit_totals = aggregate(conduits, cables)
        free_cable_totals = aggregate_free_cables(free_cables)
        bollard_totals = aggregate_bollards(bollards)
        dim_cache = _load_cache()

        st.success(f"✅ **{source_name}** を解析しました")

        # ─── 抽出テキスト（折りたたみ） ───
        with st.expander(f"抽出テキスト ({len(lines)} 行)", expanded=False):
            st.code("\n".join(lines), language=None)

        st.divider()

        # ─────────────────────────────────────────
        #  集計表: 3カラムレイアウト
        # ─────────────────────────────────────────

        col1, col2, col3 = st.columns([1, 1.4, 0.8])

        # ── 配線種 ──
        with col1:
            st.subheader("📋 配線種")
            if cable_totals:
                CABLE_ORDER = ["6.6kVCVT38sq", "CVT150sq", "IV38sq", "IV14sq", "IV5.5sq"]
                ordered = CABLE_ORDER + [k for k in cable_totals if k not in CABLE_ORDER]
                rows = [
                    {"配線種": k, "全長": f"{cable_totals[k]} m"}
                    for k in ordered if k in cable_totals
                ]
                df_cable = pd.DataFrame(rows)
                st.dataframe(df_cable, hide_index=True, use_container_width=True)
            else:
                st.warning("ケーブル情報が見つかりませんでした")

        # ── 配管種 ──
        with col2:
            st.subheader("🔧 配管種")
            if conduit_totals:
                rows = []
                empty_conduits = []
                for inst, ctype, cabs, length in conduit_totals:
                    rows.append({
                        "設置": inst,
                        "配管種": ctype,
                        "全長": f"{length} m",
                        "内線": " / ".join(cabs) if cabs else "―",
                    })
                    if not cabs:
                        empty_conduits.append(f"{inst} {ctype} {length}m")
                df_conduit = pd.DataFrame(rows)
                st.dataframe(df_conduit, hide_index=True, use_container_width=True)
                if empty_conduits:
                    st.warning(
                        "⚠️ 内線が検出されていない配管があります（要確認）:\n"
                        + "\n".join(f"- {c}" for c in empty_conduits)
                    )
            else:
                st.warning("配管情報が見つかりませんでした")

        # ── 付帯設備 ──
        with col3:
            st.subheader("🚧 付帯設備")
            accessory_rows = []

            # 盤類
            for p in panels:
                info = get_panel_info(p.model, dim_cache) if p.model else None
                detail = format_panel_detail(p.model, info)
                accessory_rows.append({"種別": p.name, "詳細": detail, "数量": "1 台"})

            # バリカー
            for label, qty in bollard_totals.items():
                accessory_rows.append({"種別": label, "詳細": "―", "数量": f"{qty} 台"})

            if accessory_rows:
                st.dataframe(pd.DataFrame(accessory_rows), hide_index=True, use_container_width=True)
            else:
                st.info("付帯設備は検出されませんでした")

        # ── 配管なしケーブル ──
        if free_cable_totals:
            st.subheader("🔌 配管なしケーブル（充電器内・ポール内・キュービクル内など）")
            rows = [
                {"設置場所": loc, "ケーブル種": ctype, "全長": f"{ln} m"}
                for loc, ctype, ln in free_cable_totals
            ]
            st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)

        st.divider()

        # ─────────────────────────────────────────
        #  Excel ダウンロード
        # ─────────────────────────────────────────

        def build_excel(cable_totals, conduit_totals, free_cable_totals, bollard_totals, panels, dim_cache) -> bytes:
            import openpyxl
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "サマリ"

            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            header_fill = PatternFill("solid", fgColor="DDEEFF")
            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")

            def cell(r, c, value, font=None, fill=None, align=None):
                cl = ws.cell(row=r, column=c, value=value)
                if font:  cl.font = font
                if fill:  cl.fill = fill
                if align: cl.alignment = align
                cl.border = border
                return cl

            row = 1

            # タイトル
            ws.merge_cells(f"A{row}:D{row}")
            c = ws.cell(row=row, column=1, value="サマリ")
            c.font = Font(bold=True, size=14)
            c.alignment = center
            row += 1

            # 配線種
            ws.merge_cells(f"A{row}:D{row}")
            cell(row, 1, "配線種", bold, header_fill, center)
            row += 1

            CABLE_ORDER = ["6.6kVCVT38sq", "CVT150sq", "IV38sq", "IV14sq", "IV5.5sq"]
            ordered = CABLE_ORDER + [k for k in cable_totals if k not in CABLE_ORDER]
            for k in ordered:
                if k not in cable_totals:
                    continue
                cell(row, 1, "", align=center)
                cell(row, 2, k, align=center)
                ws.merge_cells(f"C{row}:D{row}")
                cell(row, 3, f"全長{cable_totals[k]}m", align=center)
                ws.cell(row=row, column=4).border = border
                row += 1

            row += 1

            # 配管種
            ws.merge_cells(f"A{row}:D{row}")
            cell(row, 1, "配管種", bold, header_fill, center)
            row += 1
            for install_type in ["露出", "埋設"]:
                for inst, ctype, cabs, length in conduit_totals:
                    if inst != install_type:
                        continue
                    cable_note = " / ".join(cabs) if cabs else ""
                    cell(row, 1, inst, align=center)
                    cell(row, 2, ctype, align=center)
                    cell(row, 3, f"全長{length}m", align=center)
                    cell(row, 4, f"({cable_note})" if cable_note else "", align=left_align)
                    row += 1

            row += 1

            # 配管なしケーブル
            if free_cable_totals:
                ws.merge_cells(f"A{row}:D{row}")
                cell(row, 1, "配管なしケーブル", bold, header_fill, center)
                row += 1
                for loc, ctype, ln in free_cable_totals:
                    cell(row, 1, loc, align=center)
                    cell(row, 2, ctype, align=center)
                    ws.merge_cells(f"C{row}:D{row}")
                    cell(row, 3, f"全長{ln}m", align=center)
                    ws.cell(row=row, column=4).border = border
                    row += 1
                row += 1

            # 付帯設備
            ws.merge_cells(f"A{row}:D{row}")
            cell(row, 1, "付帯設備", bold, header_fill, center)
            row += 1
            has_accessory = bool(panels or bollard_totals)
            for p in panels:
                info = get_panel_info(p.model, dim_cache) if p.model else None
                detail = format_panel_detail(p.model, info)
                cell(row, 1, p.name, align=center)
                ws.merge_cells(f"B{row}:C{row}")
                cell(row, 2, detail, align=left_align)
                ws.cell(row=row, column=3).border = border
                cell(row, 4, "1台", align=center)
                row += 1
            for label, qty in bollard_totals.items():
                cell(row, 1, label, align=center)
                ws.merge_cells(f"B{row}:C{row}")
                cell(row, 2, "", align=center)
                ws.cell(row=row, column=3).border = border
                cell(row, 4, f"{qty}台", align=center)
                row += 1
            if not has_accessory:
                ws.merge_cells(f"A{row}:D{row}")
                cell(row, 1, "（検出されませんでした）", align=left_align)
                row += 1

            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 20
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 30

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        excel_bytes = build_excel(cable_totals, conduit_totals, free_cable_totals, bollard_totals, panels, dim_cache)
        stem = Path(source_name).stem if source_name != "直接入力" else "集計結果"

        st.download_button(
            label="📥 Excelファイルをダウンロード",
            data=excel_bytes,
            file_name=f"{stem}_サマリ.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=False,
        )
